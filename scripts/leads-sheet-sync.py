#!/usr/bin/env python3
"""
myPayAdvisor -> Google Sheet lead export.

Pulls quiz_leads + newsletter_subscribers from prod Supabase and publishes a
shareable Google Sheet owned by Assaf.

Both tables feed the Leads tab, because the sheet is the call list and a lead
that only left an email is still a lead. Email-only rows carry the priority
"6 - Email only" and blank intake columns. The Newsletter tab remains the
mailing list of record: it keeps unsubscribes, which the call list drops.

  python3 scripts/leads-sheet-sync.py create        # first run: build + upload
  python3 scripts/leads-sheet-sync.py sync          # refresh data, keep manual edits
  python3 scripts/leads-sheet-sync.py migrate-cols  # add new DB columns in place

Four tabs: Dashboard, Leads, Newsletter, Legend.

Adding a DB column later? Add it to LEAD_COLS + lead_row_values + build_leads,
then append it to COLUMN_MIGRATIONS and run `migrate-cols` once. That inserts it
into the live sheet in place, so the URL Assaf already shared keeps working and
the manual columns keep their contents. Never re-run `create` for that: it
uploads a brand new file with a brand new URL.

On the Leads tab the first LEAD_COLS columns come from the database and are
rewritten on every sync. The LEAD_EDIT_COLS block after them (Status, Owner,
Last contacted, Next follow-up, What they said) belongs to whoever is working
the sheet: sync reads it back, keys it to the lead id in the hidden last
column, and writes it to wherever that lead ended up. Formatting for
BUFFER_ROWS rows is banked at creation time so new leads land pre-styled.

Two of the DB-owned columns hold formulas rather than values ("Days open" and
"Do next"). The sync rewrites them with the right row number each run, and they
keep recalculating in between, which is why they read from the manual columns
without ever going stale.
"""

import json
import os
import re
import sys
import datetime as dt
from urllib.parse import urlencode
from urllib.error import HTTPError
from urllib.request import Request, urlopen

# openpyxl is only needed to BUILD the workbook (create mode). Sync mode talks
# to the Sheets API over plain HTTP, so the scheduled Hermes run needs no deps.
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


def col_letter(n):
    """1 -> A, 27 -> AA. Local so sync mode stays dependency-free."""
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Overridable so the same file runs from the repo on the Mac and from a flat
# deploy dir on Hermes.
ENV_FILE = os.environ.get("MPA_ENV_FILE", os.path.join(REPO, ".env.local"))
TOKEN_FILE = os.environ.get(
    "MPA_GOOGLE_TOKEN", os.path.expanduser("~/.credentials/google-sheets-token.json"))
STATE_FILE = os.environ.get("MPA_STATE_FILE",
                            os.path.join(REPO, "scripts", ".leads-sheet-state.json"))
XLSX_FILE = os.environ.get("MPA_XLSX_FILE",
                           os.path.join(REPO, "scripts", ".leads-sheet.xlsx"))
SHEET_TITLE = "myPayAdvisor — Leads"

HEADER_ROW = 4
FIRST_DATA_ROW = HEADER_ROW + 1
LEAD_BUFFER_ROWS = 400
NEWS_BUFFER_ROWS = 200

# ---------------------------------------------------------------- label maps
# Mirrored from src/lib/funnel/types.ts so the sheet reads in plain English.
BUSINESS_TYPE_LABELS = {
    "physical_goods": "Physical goods / E-commerce",
    "saas_digital": "SaaS or digital product",
    "subscription": "Membership or subscription",
    "retail_inperson": "Retail with in-person payments",
    "restaurant_hospitality": "Restaurant or hospitality",
    "field_services": "Field services or mobile payments",
    "financial_services": "Financial services or money transfers",
    "health_wellness": "Health, wellness or lifestyle products",
    "gaming_entertainment": "Gaming or entertainment",
    "other": "Other / Not sure",
}

VOLUME_TIER_LABELS = {
    "under_50k": "Under $50K / mo",
    "50k_to_250k": "$50K - $250K / mo",
    "250k_to_1m": "$250K - $1M / mo",
    "over_1m": "Over $1M / mo",
    "pre_launch": "Pre-launch / not live yet",
}

# Revenue potential for an advisory commission model. Monthly volume is the real
# signal, so it drives call order rather than the DB lead_score column.
VOLUME_PRIORITY = {
    "over_1m": ("1 - Hot", 1),
    "250k_to_1m": ("2 - High", 2),
    "50k_to_250k": ("3 - Medium", 3),
    "under_50k": ("4 - Low", 4),
    "pre_launch": ("5 - Early", 5),
}
# Its own tier rather than "4 - Low": we don't know this person's volume, we
# only know nobody asked. Blank would have dropped them out of every Dashboard
# priority count and made the totals stop adding up.
EMAIL_ONLY_PRIORITY = "6 - Email only"
PRIORITY_ORDER = ["1 - Hot", "2 - High", "3 - Medium", "4 - Low", "5 - Early",
                  EMAIL_ONLY_PRIORITY]

PAIN_POINT_LABELS = {
    "funds_frozen": "Provider is freezing funds / applying reserves",
    "approval_rates": "Approval rates underperforming, losing revenue at checkout",
    "long_onboarding": "Onboarding took months, can't repeat it",
    "new_markets": "Needs to expand into new markets",
    "failed_recurring": "Too many failed recurring payments",
    "in_person_costs": "In-person transaction costs too high",
    "needs_approval": "Launching, needs a provider that will approve them",
}

TRACK_LABELS = {
    "A": "A - Online / growth",
    "B": "B - In-person",
    "C": "C - Complex / high-risk",
    "MANUAL": "Manual",
}

SOURCE_LABELS = {
    "sorting_hat": "Sorting Hat (site quiz)",
    "quiz": "Payment quiz",
    "newsletter": "Newsletter",
    "footer": "Newsletter - footer form",
    "exit_intent": "Newsletter - exit popup",
}

# Same rows, worded for the Leads tab, where the question is "what do I know
# before I dial" rather than "which list is this".
EMAIL_ONLY_ENTRY_LABELS = {
    "exit_intent": "Exit popup - email only",
    "footer": "Footer form - email only",
    "newsletter": "Newsletter - email only",
}

# Ordered the way an outbound day actually runs, so the dropdown reads like a
# ladder rather than a bag of labels.
STATUS_OPTIONS = ["New", "Emailed", "Called - no answer", "In conversation",
                  "Call booked", "Handed to Barak", "Proposal sent",
                  "Won", "Lost", "Not a fit"]

OWNER_OPTIONS = ["Linoy", "Barak", "Assaf"]

LEADS_TITLE = "myPayAdvisor — Leads"
LEADS_SUBTITLE = (
    "Everyone who left anything on mypayadvisor.com, newest first, including "
    "the email-only signups from the exit popup and the footer form "
    "(Priority \"6 - Email only\"). Sort by \"Do next\" and work top down. The "
    "yellow columns on the right are yours to fill in and a refresh never "
    "overwrites them. Everything else is rewritten every morning, so don't type "
    "there. Full guide on the Legend tab."
)

TRAFFIC_BUCKETS = ["ChatGPT (AI search)", "Google (organic)", "Perplexity (AI search)",
                   "Gemini (AI search)", "Claude (AI search)", "Direct / unknown"]

# --------------------------------------------------------------- style tokens
NAVY, SLATE, LINE, BAND = "0F2740", "44566C", "D8DEE6", "F5F8FB"
EDIT_BG, HOT, WARM, COOL = "FFF8E6", "FDE7E7", "FFF3E0", "EAF4EC"

if HAVE_OPENPYXL:
    HEAD_FILL = PatternFill("solid", fgColor=NAVY)
    EDIT_HEAD_FILL = PatternFill("solid", fgColor="8A6D1F")
    EDIT_FILL = PatternFill("solid", fgColor=EDIT_BG)
    HEAD_FONT = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Inter", size=10, color="1B2733")
    MUTED_FONT = Font(name="Inter", size=10, color=SLATE)
    TITLE_FONT = Font(name="Inter", size=16, bold=True, color=NAVY)
    THIN = Side(style="thin", color=LINE)
    CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ------------------------------------------------------------------- plumbing
def read_env(path):
    env = {}
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def supabase_get(url, key, table, query):
    req = Request(f"{url}/rest/v1/{table}?{query}",
                  headers={"apikey": key, "Authorization": f"Bearer {key}"})
    with urlopen(req) as resp:
        return json.loads(resp.read())


def google_access_token():
    with open(TOKEN_FILE) as fh:
        tok = json.load(fh)
    body = urlencode({
        "client_id": tok["client_id"], "client_secret": tok["client_secret"],
        "refresh_token": tok["refresh_token"], "grant_type": "refresh_token",
    }).encode()
    req = Request(tok["token_uri"], data=body,
                  headers={"Content-Type": "application/x-www-form-urlencoded"})
    with urlopen(req) as resp:
        return json.loads(resp.read())["access_token"]


def google_json(token, url, method="GET", payload=None):
    data = json.dumps(payload).encode() if payload is not None else None
    req = Request(url, data=data, method=method, headers={
        "Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    try:
        with urlopen(req) as resp:
            raw = resp.read()
    except HTTPError as err:
        # The body is where Google says what it actually objected to, and a
        # bare "HTTP Error 400" tells you nothing.
        detail = err.read().decode("utf-8", "replace")[:1000]
        raise SystemExit(f"Google API {err.code} on {method} {url}\n{detail}")
    return json.loads(raw) if raw else {}


# ------------------------------------------------------------- data shaping
def traffic_source(row):
    """Where the lead actually came from, in words."""
    blob = " ".join([(row.get("utm_source") or ""), (row.get("referrer") or ""),
                     (row.get("landing_page_url") or "")]).lower()
    for needle, label in [("chatgpt", "ChatGPT (AI search)"),
                          ("perplexity", "Perplexity (AI search)"),
                          ("gemini", "Gemini (AI search)"),
                          ("bard", "Gemini (AI search)"),
                          ("claude", "Claude (AI search)"),
                          ("google", "Google (organic)"),
                          ("bing", "Bing"), ("linkedin", "LinkedIn")]:
        if needle in blob:
            return label
    ref = row.get("referrer") or ""
    if ref:
        return re.sub(r"^https?://(www\.)?", "", ref).split("/")[0] or "Referral"
    return "Direct / unknown"


def parse_ts(value):
    """Postgres timestamps carry 1-6 fractional digits; fromisoformat on
    Python < 3.11 only accepts exactly 3 or 6, so normalise to 6 first."""
    if not value:
        return None
    v = value.replace("Z", "+00:00")
    v = re.sub(r"\.(\d{1,6})(?=[+-]\d{2}:?\d{2}$|$)",
               lambda m: "." + m.group(1).ljust(6, "0"), v)
    try:
        parsed = dt.datetime.fromisoformat(v)
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(dt.timezone.utc).replace(tzinfo=None)
    return parsed


def build_leads(rows):
    """quiz_leads rows -> display rows, newest first.

    Duplicate flagging happens after the merge in build_lead_stream(), not here,
    so that someone who took the quiz *and* left their email in the popup reads
    as one person rather than two leads."""
    out = []
    for row in sorted(rows, key=lambda r: r.get("created_at") or "", reverse=True):
        vol = row.get("volume_tier") or row.get("monthly_volume") or ""
        priority, _ = VOLUME_PRIORITY.get(vol, ("4 - Low", 4))
        utm = " · ".join(v for v in (row.get("utm_source"), row.get("utm_medium"),
                                     row.get("utm_campaign")) if v)
        entry = row.get("source") or row.get("lead_source") or ""
        out.append({
            "id": row["id"],
            "received": parse_ts(row.get("created_at")),
            "name": row.get("full_name") or "",
            "company": row.get("company_name") or "",
            "email": row.get("email") or "",
            "phone": row.get("phone") or "",
            "provider": row.get("current_provider") or "",
            "business": BUSINESS_TYPE_LABELS.get(row.get("business_type") or "",
                                                 row.get("industry") or ""),
            "volume": VOLUME_TIER_LABELS.get(vol, vol),
            "priority": priority,
            "pain": PAIN_POINT_LABELS.get(row.get("pain_point") or "", ""),
            "track": TRACK_LABELS.get(row.get("track") or "", row.get("track") or ""),
            "came_from": traffic_source(row),
            "entry": SOURCE_LABELS.get(entry, entry),
            "landing": (row.get("landing_page_url") or "").split("?")[0],
            # Raw attribution, kept alongside the readable "Came from" so
            # nothing the lead arrived with is hidden.
            "referrer": row.get("referrer") or "",
            "utm": utm,
            "enriched": parse_ts(row.get("enriched_at")),
            "dup": "",
        })
    return out


def build_email_only_leads(rows):
    """newsletter_subscribers rows -> the same display shape as build_leads.

    These people raised a hand on the site and were only ever asked for an
    email, so the intake columns are legitimately blank. They still belong on
    the Leads tab: the sheet is the call list, and an unworked row is the only
    thing that costs money. Unsubscribes are left out; chasing someone who
    opted out is how a domain gets burned.

    Phone / Company / Processing with today arrive only if they filled the
    optional step after signing up, exactly like step 5 on the quiz side, so
    blank there is normal and means they skipped it.

    Attribution comes from the same five columns and the same first-touch
    cookie as quiz_leads, so traffic_source() reads a row from either table
    unchanged. Signups written before 2026-08-12 have none of it: the cookie
    was being collected and then discarded on this path, and that history
    cannot be reconstructed, so they stay "Direct / unknown".
    """
    out = []
    for row in sorted(rows, key=lambda r: r.get("subscribed_at") or "",
                      reverse=True):
        if row.get("active") is False:
            continue
        src = row.get("source") or ""
        utm = " · ".join(v for v in (row.get("utm_source"), row.get("utm_medium"),
                                     row.get("utm_campaign")) if v)
        out.append({
            "id": row["id"],
            "received": parse_ts(row.get("subscribed_at")),
            "name": "",
            "business": "", "volume": "", "pain": "", "track": "",
            "email": row.get("email") or "",
            # Populated by the optional second step of the popup.
            "phone": row.get("phone") or "",
            "company": row.get("company_name") or "",
            "provider": row.get("current_provider") or "",
            "enriched": parse_ts(row.get("enriched_at")),
            "priority": EMAIL_ONLY_PRIORITY,
            "came_from": traffic_source(row),
            "entry": EMAIL_ONLY_ENTRY_LABELS.get(
                src, SOURCE_LABELS.get(src, src or "Email only")),
            "landing": (row.get("landing_page_url") or "").split("?")[0],
            "referrer": row.get("referrer") or "",
            "utm": utm,
            "dup": "",
        })
    return out


def build_lead_stream(quiz_rows, subscriber_rows):
    """One call list, newest first, from both capture paths.

    Dedup runs across the union on email: the newest row stays clean and older
    ones are flagged, so a quiz lead who had already left their email in the
    popup shows up once as a real lead rather than twice.
    """
    merged = build_leads(quiz_rows) + build_email_only_leads(subscriber_rows)
    merged.sort(key=lambda item: (item["received"] or dt.datetime.min),
                reverse=True)
    seen = set()
    for item in merged:
        email = (item["email"] or "").strip().lower()
        if not email:
            continue
        item["dup"] = "earlier dup" if email in seen else ""
        seen.add(email)
    return merged


def build_newsletter(rows):
    return [{
        "id": row["id"],
        "subscribed": parse_ts(row.get("subscribed_at")),
        "email": row.get("email") or "",
        "source": SOURCE_LABELS.get(row.get("source") or "",
                                    row.get("source") or "Newsletter"),
        "active": "Yes" if row.get("active") is not False else "Unsubscribed",
    } for row in sorted(rows, key=lambda r: r.get("subscribed_at") or "", reverse=True)]


def lead_row_values(lead, row):
    return [lead["received"], days_open_formula(row), do_next_formula(row),
            lead["name"], lead["company"], lead["email"], lead["phone"],
            lead["business"], lead["volume"], lead["priority"], lead["pain"],
            lead["provider"], lead["track"], lead["came_from"], lead["entry"],
            lead["landing"], lead["referrer"], lead["utm"], lead["enriched"],
            lead["dup"]]


def news_row_values(sub, row=None):
    return [sub["subscribed"], sub["email"], sub["source"], sub["active"]]


# ------------------------------------------------------------ sheet building
# Everything the lead actually gave us, left to right in the order you'd want
# it before picking up the phone. Company / Phone / Processing with today come
# from the optional step 5, so they are blank for anyone who skipped it.
# "Days open" and "Do next" are live formulas, not stored data.
LEAD_COLS = [("Received", 17), ("Days open", 11), ("Do next", 18),
             ("Name", 16), ("Company", 22), ("Email", 30), ("Phone", 16),
             ("Business", 30), ("Monthly volume", 19), ("Priority", 12),
             ("What they need help with", 46), ("Processing with today", 22),
             ("Track", 22), ("Came from", 20), ("Entry point", 20),
             ("Landing page", 40), ("Referrer", 30), ("UTM", 24),
             ("Details added", 15), ("Flag", 11)]
LEAD_EDIT_COLS = [("Status", 18), ("Owner", 13), ("Last contacted", 16),
                  ("Next follow-up", 16), ("What they said", 46)]
NEWS_TITLE = "Newsletter — mailing list"
NEWS_SUBTITLE = ("The mailing list, including anyone who unsubscribed. Everyone "
                 "still active also appears on the Leads tab as \"6 - Email "
                 "only\" — work them there, this tab is for sending, not calling.")

NEWS_COLS = [("Subscribed", 17), ("Email", 34), ("Signed up at", 24), ("Active", 14)]
NEWS_EDIT_COLS = [("Status", 16), ("Notes", 40)]

N_LEAD_DATA = len(LEAD_COLS)                       # A..N
N_LEAD_TOTAL = N_LEAD_DATA + len(LEAD_EDIT_COLS)   # ..R
LEAD_ID_COL = N_LEAD_TOTAL + 1                     # S, hidden
N_NEWS_DATA = len(NEWS_COLS)
N_NEWS_TOTAL = N_NEWS_DATA + len(NEWS_EDIT_COLS)
NEWS_ID_COL = N_NEWS_TOTAL + 1


ALL_LEAD_HEADERS = [c[0] for c in LEAD_COLS] + [c[0] for c in LEAD_EDIT_COLS]


def lead_col(label):
    """Column letter for any Leads header, data or manual. Every formula goes
    through this, so moving a column can never silently mis-point one."""
    return col_letter(1 + ALL_LEAD_HEADERS.index(label))


LEAD_STATUS_COL = col_letter(N_LEAD_DATA + 1)


def days_open_formula(row):
    """How long this lead has been sitting, recalculated every time the sheet
    is opened rather than frozen at the last sync."""
    return f'=IF($A{row}="","",INT(TODAY()-$A{row}))'


def do_next_formula(row):
    """The one column to work from: turns status, last contact and the
    follow-up date into a single instruction. Written by the sync as a live
    formula, so it keeps answering correctly between runs."""
    st, last = lead_col("Status"), lead_col("Last contacted")
    nxt = lead_col("Next follow-up")
    return (
        f'=IF($A{row}="","",'
        f'IF(OR(${st}{row}="Won",${st}{row}="Lost",${st}{row}="Not a fit"),"Closed",'
        f'IF(AND(${nxt}{row}<>"",${nxt}{row}<=TODAY()),"Follow up due",'
        f'IF(OR(${st}{row}="",${st}{row}="New"),"Contact now",'
        f'IF(${nxt}{row}<>"","Booked "&TEXT(${nxt}{row},"dd/mm"),'
        f'IF(AND(${last}{row}<>"",TODAY()-${last}{row}>=4),'
        f'"Chase, "&INT(TODAY()-${last}{row})&"d silent","Working"))))))'
    )

# Columns that only the optional step 5 can fill.
OPTIONAL_LEAD_COLS = ["Company", "Phone", "Processing with today"]


def extra_details_formula(lo, hi):
    """How many leads went past the required fields and told us more."""
    any_of = "+".join(
        f'(Leads!${lead_col(c)}${lo}:${lead_col(c)}${hi}<>"")'
        for c in OPTIONAL_LEAD_COLS)
    em = lead_col("Email")
    return f'=SUMPRODUCT((Leads!${em}${lo}:${em}${hi}<>"")*(({any_of})>0))'


def style_header(ws, cols, edit_cols, id_col):
    for i, (label, width) in enumerate(cols + edit_cols, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=label)
        cell.fill = EDIT_HEAD_FILL if i > len(cols) else HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
        ws.column_dimensions[col_letter(i)].width = width
    idc = ws.cell(row=HEADER_ROW, column=id_col, value="id")
    idc.fill = HEAD_FILL
    idc.font = HEAD_FONT
    ws.column_dimensions[col_letter(id_col)].hidden = True
    ws.row_dimensions[HEADER_ROW].height = 30


def style_body_cell(ws, r, c, n_data, total, wrap_cols, date_cols, id_col):
    cell = ws.cell(row=r, column=c)
    cell.font = BODY_FONT
    cell.border = CELL_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=(c in wrap_cols))
    if c > n_data and c <= total:
        cell.fill = EDIT_FILL
    if c in date_cols:
        cell.number_format = "yyyy-mm-dd hh:mm" if c == 1 else "yyyy-mm-dd"
    return cell


def write_leads_tab(wb, leads):
    ws = wb.create_sheet("Leads")
    ws.sheet_properties.tabColor = NAVY
    ws["A1"] = LEADS_TITLE
    ws["A1"].font = TITLE_FONT
    ws["A2"] = LEADS_SUBTITLE
    ws["A2"].font = MUTED_FONT
    # Deliberately not merged: a merge spanning the frozen-column boundary
    # makes Sheets refuse to freeze at all.
    ws.row_dimensions[2].height = 18

    style_header(ws, LEAD_COLS, LEAD_EDIT_COLS, LEAD_ID_COL)

    wrap_cols = {LEAD_COLS.index(c) + 1 for c in LEAD_COLS
                 if c[0] in ("Business", "What they need help with", "Landing page")}
    wrap_cols.add(N_LEAD_TOTAL)
    date_cols = {ALL_LEAD_HEADERS.index(h) + 1 for h in
                 ("Received", "Details added", "Last contacted", "Next follow-up")}
    for i in range(LEAD_BUFFER_ROWS):
        r = FIRST_DATA_ROW + i
        lead = leads[i] if i < len(leads) else None
        values = lead_row_values(lead, r) if lead else [None] * N_LEAD_DATA
        for c in range(1, N_LEAD_TOTAL + 1):
            cell = style_body_cell(ws, r, c, N_LEAD_DATA, N_LEAD_TOTAL,
                                   wrap_cols, date_cols, LEAD_ID_COL)
            if c <= N_LEAD_DATA:
                cell.value = values[c - 1]
            elif c == N_LEAD_DATA + 1 and lead:
                cell.value = "New"
        if lead:
            ws.cell(row=r, column=LEAD_ID_COL, value=lead["id"])
        ws.row_dimensions[r].height = 30

    last = FIRST_DATA_ROW + LEAD_BUFFER_ROWS - 1
    ws.auto_filter.ref = f"A{HEADER_ROW}:{col_letter(N_LEAD_TOTAL)}{last}"

    pc = lead_col("Priority")
    prio = f"{pc}{FIRST_DATA_ROW}:{pc}{last}"
    ws.conditional_formatting.add(prio, CellIsRule(
        operator="equal", formula=['"1 - Hot"'], fill=PatternFill("solid", fgColor=HOT),
        font=Font(bold=True, color="9B1C1C")))
    ws.conditional_formatting.add(prio, CellIsRule(
        operator="equal", formula=['"2 - High"'], fill=PatternFill("solid", fgColor=WARM),
        font=Font(bold=True, color="8A4B00")))

    status_range = f"{LEAD_STATUS_COL}{FIRST_DATA_ROW}:{LEAD_STATUS_COL}{last}"
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Won"'], fill=PatternFill("solid", fgColor=COOL),
        font=Font(bold=True, color="1B5E20")))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"Lost"'], font=Font(color="9AA5B1", italic=True)))
    ws.conditional_formatting.add(
        f"A{FIRST_DATA_ROW}:{col_letter(N_LEAD_DATA)}{last}",
        FormulaRule(formula=[f'${lead_col("Flag")}{FIRST_DATA_ROW}="earlier dup"'],
                    font=Font(color="9AA5B1", italic=True)))

    # Same colour language as refresh_leads_rules applies to the live sheet.
    dnc = lead_col("Do next")
    do_next_range = f"{dnc}{FIRST_DATA_ROW}:{dnc}{last}"
    for value, fill, colour, bold in (
            ("Contact now", HOT, "9B1C1C", True),
            ("Follow up due", WARM, "8A4B00", True),
            ("Closed", "F7F7F7", "99A3AD", False)):
        ws.conditional_formatting.add(do_next_range, CellIsRule(
            operator="equal", formula=[f'"{value}"'],
            fill=PatternFill("solid", fgColor=fill),
            font=Font(bold=bold, color=colour, italic=value == "Closed")))
    ws.conditional_formatting.add(do_next_range, FormulaRule(
        formula=[f'LEFT({dnc}{FIRST_DATA_ROW},5)="Chase"'],
        fill=PatternFill("solid", fgColor=WARM), font=Font(color="8A4B00")))

    doc = lead_col("Days open")
    ws.conditional_formatting.add(
        f"{doc}{FIRST_DATA_ROW}:{doc}{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["7"],
                   fill=PatternFill("solid", fgColor=HOT),
                   font=Font(bold=True, color="9B1C1C")))

    for options, rng in ((STATUS_OPTIONS, status_range),
                         (OWNER_OPTIONS,
                          f"{lead_col('Owner')}{FIRST_DATA_ROW}:"
                          f"{lead_col('Owner')}{last}")):
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                            allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(rng)

    # Keep Received, Days open, Do next and Name on screen while scrolling right.
    ws.freeze_panes = f"{lead_col('Company')}{FIRST_DATA_ROW}"


def write_newsletter_tab(wb, subs):
    ws = wb.create_sheet("Newsletter")
    ws.sheet_properties.tabColor = "3E6E8E"
    ws["A1"] = NEWS_TITLE
    ws["A1"].font = TITLE_FONT
    ws["A2"] = NEWS_SUBTITLE
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A2:D2")

    style_header(ws, NEWS_COLS, NEWS_EDIT_COLS, NEWS_ID_COL)

    for i in range(NEWS_BUFFER_ROWS):
        r = FIRST_DATA_ROW + i
        sub = subs[i] if i < len(subs) else None
        values = news_row_values(sub, r) if sub else [None] * N_NEWS_DATA
        for c in range(1, N_NEWS_TOTAL + 1):
            cell = style_body_cell(ws, r, c, N_NEWS_DATA, N_NEWS_TOTAL,
                                   {N_NEWS_TOTAL}, {1}, NEWS_ID_COL)
            if c <= N_NEWS_DATA:
                cell.value = values[c - 1]
            elif c == N_NEWS_DATA + 1 and sub:
                cell.value = "New"
        if sub:
            ws.cell(row=r, column=NEWS_ID_COL, value=sub["id"])
        ws.row_dimensions[r].height = 22

    last = FIRST_DATA_ROW + NEWS_BUFFER_ROWS - 1
    ws.auto_filter.ref = f"A{HEADER_ROW}:{col_letter(N_NEWS_TOTAL)}{last}"
    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def write_dashboard_tab(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = "1B5E20"
    ws["A1"] = "Overview"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Live formulas. Updates itself as the Leads tab changes."
    ws["A2"].font = MUTED_FONT
    for col, width in [("A", 34), ("B", 14), ("C", 4), ("D", 34), ("E", 14)]:
        ws.column_dimensions[col].width = width

    def block(lc, vc, start, title, pairs):
        head = ws[f"{lc}{start}"]
        head.value = title
        head.font = Font(name="Inter", size=11, bold=True, color=NAVY)
        head.fill = PatternFill("solid", fgColor=BAND)
        ws[f"{vc}{start}"].fill = PatternFill("solid", fgColor=BAND)
        for i, (label, formula) in enumerate(pairs, start=start + 1):
            ws[f"{lc}{i}"] = label
            ws[f"{lc}{i}"].font = BODY_FONT
            cell = ws[f"{vc}{i}"]
            cell.value = formula
            cell.font = Font(name="Inter", size=10, bold=True, color=NAVY)
            cell.alignment = Alignment(horizontal="right")

    # Single source of truth, shared with the migration path so a rebuilt sheet
    # and a migrated one can never disagree.
    formula = dashboard_formulas()
    for lc, vc, start, title, labels in dashboard_layout():
        block(lc, vc, start, title, [(l, formula[l]) for l in labels])


LEGEND_ROWS = [
    ("", ""),
    ("SECTION", "Start here"),
    ("The short version", "Sort or filter the Leads tab by 'Do next'. Work "
                          "'Contact now' first, then 'Follow up due'. Set Status "
                          "and Next follow-up when you're done with a lead, and "
                          "the sheet tells you the rest by itself."),
    ("Only the yellow columns", "Everything white is written by the site and gets "
                                "overwritten every morning. Typing there will be "
                                "lost. The yellow block on the right is yours and "
                                "is never touched."),
    ("", ""),
    ("SECTION", "Tabs"),
    ("Leads", "The call list. Everyone who left anything on the site, newest "
              "first: full intakes and email-only signups together."),
    ("Newsletter", "The mailing list, unsubscribes included. For sending, not "
                   "for calling. Everyone active on it is already on Leads."),
    ("Dashboard", "Counts itself. 'Work queue' at the top is today's list."),
    ("", ""),
    ("SECTION", "Do next — the column to work from"),
    ("Contact now", "Nobody has spoken to them yet. Red."),
    ("Follow up due", "You set a Next follow-up date and it has arrived. Amber."),
    ("Chase, Nd silent", "You reached out N days ago and nothing came back."),
    ("Booked dd/mm", "A follow-up is set for the future. Nothing to do today."),
    ("Working", "In conversation, no date set. Set one so it stops being "
                "invisible."),
    ("Closed", "Won, Lost or Not a fit. Greyed out."),
    ("", ""),
    ("SECTION", "What the lead told us"),
    ("Business / Monthly volume / What they need help with",
     "The intake answers. Anyone who came through the quiz has these, so never "
     "open with 'what do you need' — they already said. Blank across all three "
     "means a '6 - Email only' row, where asking is exactly the right opener."),
    ("Company / Phone / Processing with today",
     "The optional step after signup. Blank is normal and means they skipped "
     "it. When Company is blank the email domain is usually the business, and "
     "a phone number means they'd rather be called than written to."),
    ("Came from / Entry point / Landing page / Referrer / UTM",
     "How they found us and what they were reading when they raised a hand. "
     "A ChatGPT or Perplexity lead found us through an AI answer, which usually "
     "means they've been comparing for a while."),
    ("Details added", "Date they filled the optional step. Blank = skipped it."),
    ("Days open", "Days since they came in. Red past a week."),
    ("", ""),
    ("SECTION", "Priority — who to call first"),
    ("1 - Hot", "Over $1M monthly volume. Biggest residual if they switch."),
    ("2 - High", "$250K - $1M monthly."),
    ("3 - Medium", "$50K - $250K monthly."),
    ("4 - Low", "Under $50K monthly."),
    ("5 - Early", "Pre-launch, not processing yet. Long game."),
    (EMAIL_ONLY_PRIORITY,
     "Left an email in the exit popup or the footer form and was never asked "
     "anything else, so Business, Volume and Track are blank. Unknown, not "
     "small: nobody has qualified them yet. Open by asking what they process."),
    ("", ""),
    ("SECTION", "Track — which funnel routed them"),
    (TRACK_LABELS["A"], "Online, e-commerce or SaaS. Standard growth advisory."),
    (TRACK_LABELS["B"], "In-person: retail, restaurant, field services. "
                        "Hardware and per-location costs matter most."),
    (TRACK_LABELS["C"], "Complex or high-risk: financial, health, gaming, or "
                        "anyone with frozen funds, approval or onboarding "
                        "problems. Usually the highest advisory value."),
    ("", ""),
    ("SECTION", "Columns you fill in (yellow)"),
    ("Status", "Dropdown, in the order a deal actually moves. Drives the "
               "Dashboard counts."),
    ("Owner", "Linoy, Barak or Assaf."),
    ("Last contacted", "Date you last reached out. Feeds the 'Chase' warning."),
    ("Next follow-up", "Date to come back to them. This is what puts a lead "
                       "back in the work queue, so always set one."),
    ("What they said", "Anything worth remembering before the next call."),
    ("", ""),
    ("SECTION", "Housekeeping"),
    ("Flag = earlier dup", "Same email submitted more than once. The earlier "
                           "attempts are greyed out. The clean row at the top "
                           "is the current one — work that."),
    ("Refresh", "Runs every morning at 08:05. New leads land at the top and "
                "your yellow columns follow their lead to its new row."),
]


def write_legend_tab(wb):
    ws = wb.create_sheet("Legend")
    ws.sheet_properties.tabColor = SLATE
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 86
    ws["A1"] = "How to read this sheet"
    ws["A1"].font = TITLE_FONT

    for i, (label, text) in enumerate(LEGEND_ROWS, start=2):
        if label == "SECTION":
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
            cell.fill = HEAD_FILL
            ws.cell(row=i, column=2).fill = HEAD_FILL
            continue
        a = ws.cell(row=i, column=1, value=label)
        a.font = Font(name="Inter", size=10, bold=True, color=NAVY)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws.cell(row=i, column=2, value=text)
        b.font = BODY_FONT
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if text and len(text) > 80:
            ws.row_dimensions[i].height = 30


def build_workbook(leads, subs, path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_leads_tab(wb, leads)
    write_newsletter_tab(wb, subs)
    write_dashboard_tab(wb)
    write_legend_tab(wb)
    wb.save(path)
    return path


# --------------------------------------------------------------- drive/sheets
def upload_new(token, path):
    meta = {"name": SHEET_TITLE, "mimeType": "application/vnd.google-apps.spreadsheet"}
    boundary = "==mpa-leads-boundary=="
    with open(path, "rb") as fh:
        blob = fh.read()
    body = b"".join([
        f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n".encode(),
        json.dumps(meta).encode(), b"\r\n",
        f"--{boundary}\r\nContent-Type: application/vnd.openxmlformats-"
        f"officedocument.spreadsheetml.sheet\r\n\r\n".encode(),
        blob, f"\r\n--{boundary}--\r\n".encode(),
    ])
    req = Request("https://www.googleapis.com/upload/drive/v3/files"
                  "?uploadType=multipart&fields=id,webViewLink",
                  data=body, method="POST",
                  headers={"Authorization": f"Bearer {token}",
                           "Content-Type": f"multipart/related; boundary={boundary}"})
    with urlopen(req) as resp:
        return json.loads(resp.read())


def fmt_cell(value):
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return value


def sync_tab(token, sheet_id, tab, items, row_values, n_data, n_total, id_col,
             buffer_rows):
    """Rewrite DB columns, carry manual columns across by row id."""
    end_col = col_letter(id_col)
    last = FIRST_DATA_ROW + buffer_rows - 1
    rng = f"{tab}!A{FIRST_DATA_ROW}:{end_col}{last}"
    existing = google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{rng}"
    ).get("values", [])

    # id -> manual column values as currently typed in the sheet
    manual = {}
    for row in existing:
        padded = row + [""] * (id_col - len(row))
        rid = padded[id_col - 1]
        if rid:
            manual[rid] = padded[n_data:n_total]

    grid = []
    for item in items:
        kept = manual.get(item["id"], [])
        kept = (kept + [""] * (n_total - n_data))[: n_total - n_data]
        if not kept[0]:
            kept[0] = "New"
        grid.append([fmt_cell(v) for v in row_values(item, FIRST_DATA_ROW + len(grid))]
                    + kept + [item["id"]])
    # blank out the tail so deleted rows do not linger
    grid += [[""] * id_col for _ in range(buffer_rows - len(items))]

    google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{rng}"
        "?valueInputOption=USER_ENTERED",
        method="PUT", payload={"range": rng, "majorDimension": "ROWS", "values": grid},
    )
    return len(items)


# ---------------------------------------------------------- column migration
# Columns added after the sheet was first created. Position comes from the
# label's index in ALL_LEAD_HEADERS, so entries stay declarative; only the
# labels matter and order in this list is irrelevant.
#
# Inserts are applied LOW index first. That is the direction that works: after
# inserting a column, every higher target index becomes valid, whereas going
# high-to-low would aim at positions the earlier inserts have not created yet.
COLUMN_MIGRATIONS = [
    {"label": "Company"},
    {"label": "Processing with today"},
    {"label": "Days open", "number_format": "0"},
    {"label": "Do next"},
    {"label": "Entry point"},
    {"label": "Referrer"},
    {"label": "UTM"},
    {"label": "Details added", "number_format": "yyyy-mm-dd"},
    {"label": "Next follow-up", "number_format": "yyyy-mm-dd"},
]


def tab_id(token, sheet_id, title):
    meta = google_json(
        token, f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
               "?fields=sheets.properties")
    for sheet in meta.get("sheets", []):
        props = sheet.get("properties", {})
        if props.get("title") == title:
            return props["sheetId"]
    raise SystemExit(f"tab {title!r} not found in the spreadsheet")


def migrate_columns(token, sheet_id):
    """Insert any missing LEAD_COLS into the live sheet, in place.

    Inserting (rather than rebuilding) is the whole point: the spreadsheet keeps
    its URL, its manual columns keep their contents, and Google shifts the
    conditional formatting, data validation and Dashboard formulas along with
    the data. `inheritFromBefore` makes the new column pick up the borders,
    fills and fonts of its left-hand neighbour, so it lands pre-styled.
    """
    leads_id = tab_id(token, sheet_id, "Leads")
    header_rng = f"Leads!A{HEADER_ROW}:{col_letter(LEAD_ID_COL)}{HEADER_ROW}"
    header = google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{header_rng}"
    ).get("values", [[]])
    present = set(header[0] if header else [])

    pending = [m for m in COLUMN_MIGRATIONS if m["label"] not in present]
    if not pending:
        print("sheet columns already current, nothing to migrate")
        return False

    widths = dict(LEAD_COLS + LEAD_EDIT_COLS)
    pending.sort(key=lambda m: ALL_LEAD_HEADERS.index(m["label"]))

    requests = []
    for m in pending:
        at = ALL_LEAD_HEADERS.index(m["label"])
        span = {"sheetId": leads_id, "dimension": "COLUMNS",
                "startIndex": at, "endIndex": at + 1}
        requests.append({"insertDimension": {"range": span,
                                             "inheritFromBefore": True}})
        requests.append({"updateDimensionProperties": {
            "range": span,
            "properties": {"pixelSize": widths[m["label"]] * 7 + 5},
            "fields": "pixelSize"}})
        # inheritFromBefore copies the left neighbour's number format too, which
        # is wrong for a count next to a timestamp. Set it explicitly.
        requests.append({"repeatCell": {
            "range": {"sheetId": leads_id, "startRowIndex": FIRST_DATA_ROW - 1,
                      "endRowIndex": FIRST_DATA_ROW - 1 + LEAD_BUFFER_ROWS,
                      "startColumnIndex": at, "endColumnIndex": at + 1},
            "cell": {"userEnteredFormat": {"numberFormat": (
                {"type": "DATE" if "y" in m.get("number_format", "") else "NUMBER",
                 "pattern": m["number_format"]}
                if m.get("number_format") else {"type": "TEXT"})}},
            "fields": "userEnteredFormat.numberFormat"}})

    google_json(token,
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}:batchUpdate",
                method="POST", payload={"requests": requests})

    # Rewrite the whole header row so every label sits over the right column.
    google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{header_rng}"
        "?valueInputOption=RAW", method="PUT",
        payload={"range": header_rng, "majorDimension": "ROWS",
                 "values": [[c[0] for c in LEAD_COLS] +
                            [c[0] for c in LEAD_EDIT_COLS] + ["id"]]})

    print(f"inserted {len(pending)} column(s): "
          + ", ".join(m["label"] for m in pending))
    return True


def rebuild_legend(token, sheet_id):
    """Rewrite the Legend from LEGEND_ROWS. Same reasoning as the Dashboard:
    nothing here is user data, so replacing beats patching."""
    # A section row carries its title in the second slot; everything else is a
    # plain label/description pair.
    values = [["How to read this sheet", ""]]
    values += [[row[1], ""] if row[0] == "SECTION" else list(row)
               for row in LEGEND_ROWS]
    height = max(len(values), 70)
    rng = f"Legend!A1:B{height}"
    padded = values + [["", ""]] * (height - len(values))
    google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{rng}"
        "?valueInputOption=RAW", method="PUT",
        payload={"range": rng, "majorDimension": "ROWS", "values": padded})

    legend_id = tab_id(token, sheet_id, "Legend")
    requests = [{"updateDimensionProperties": {
        "range": {"sheetId": legend_id, "dimension": "COLUMNS",
                  "startIndex": i, "endIndex": i + 1},
        "properties": {"pixelSize": w}, "fields": "pixelSize"}}
        for i, w in ((0, 240), (1, 620))]
    # Reset every row, then re-bold the section headers.
    requests.append({"repeatCell": {
        "range": {"sheetId": legend_id, "startRowIndex": 1,
                  "endRowIndex": len(padded), "startColumnIndex": 0,
                  "endColumnIndex": 2},
        "cell": {"userEnteredFormat": {
            "backgroundColor": {"red": 1, "green": 1, "blue": 1},
            "wrapStrategy": "WRAP",
            "verticalAlignment": "TOP",
            "textFormat": {"bold": False, "foregroundColor": {
                "red": 0.11, "green": 0.15, "blue": 0.20}}}},
        "fields": "userEnteredFormat(backgroundColor,wrapStrategy,"
                  "verticalAlignment,textFormat)"}})
    for i, row in enumerate(LEGEND_ROWS, start=1):
        if row[0] != "SECTION":
            continue
        requests.append({"repeatCell": {
            "range": {"sheetId": legend_id, "startRowIndex": i,
                      "endRowIndex": i + 1, "startColumnIndex": 0,
                      "endColumnIndex": 2},
            "cell": {"userEnteredFormat": {
                "backgroundColor": {"red": 0.06, "green": 0.15, "blue": 0.25},
                "textFormat": {"bold": True, "foregroundColor": {
                    "red": 1, "green": 1, "blue": 1}}}},
            "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
    google_json(token,
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}:batchUpdate",
                method="POST", payload={"requests": requests})
    print(f"legend: rewrote {len(LEGEND_ROWS)} rows")


def refresh_newsletter_header(token, sheet_id):
    """Rewrite the Newsletter title and subtitle on the live sheet.

    sync_tab only ever touches data rows, so without this the tab keeps
    whatever wording it was created with, and the wording is what tells whoever
    opens it that the calling happens on Leads."""
    google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/"
        "Newsletter!A1:A2?valueInputOption=RAW", method="PUT",
        payload={"range": "Newsletter!A1:A2", "majorDimension": "ROWS",
                 "values": [[NEWS_TITLE], [NEWS_SUBTITLE]]})
    print("newsletter tab: header rewritten")


def refresh_leads_rules(token, sheet_id):
    """Re-apply the dropdowns and the colour coding on the live Leads tab.

    Replaces rather than appends: every rule this sheet should have is defined
    here, so running it twice leaves exactly one copy of each.
    """
    leads_id = tab_id(token, sheet_id, "Leads")
    first, last = FIRST_DATA_ROW - 1, FIRST_DATA_ROW - 1 + LEAD_BUFFER_ROWS

    def span(label):
        at = ALL_LEAD_HEADERS.index(label)
        return {"sheetId": leads_id, "startRowIndex": first, "endRowIndex": last,
                "startColumnIndex": at, "endColumnIndex": at + 1}

    meta = google_json(
        token, f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
               "?fields=sheets(properties.sheetId,conditionalFormats)")
    existing = 0
    for sheet in meta.get("sheets", []):
        if sheet.get("properties", {}).get("sheetId") == leads_id:
            existing = len(sheet.get("conditionalFormats", []))

    requests = [{"deleteConditionalFormatRule": {"sheetId": leads_id, "index": i}}
                for i in range(existing - 1, -1, -1)]

    for label, options in (("Status", STATUS_OPTIONS), ("Owner", OWNER_OPTIONS)):
        requests.append({"setDataValidation": {
            "range": span(label),
            "rule": {"condition": {
                "type": "ONE_OF_LIST",
                "values": [{"userEnteredValue": o} for o in options]},
                "showCustomUi": True, "strict": False}}})

    def rule(target, condition, values, bg, fg, bold=False, italic=False):
        requests.append({"addConditionalFormatRule": {"index": 0, "rule": {
            "ranges": [span(target)],
            "booleanRule": {
                "condition": {"type": condition,
                              "values": [{"userEnteredValue": v} for v in values]},
                "format": {"backgroundColor": bg,
                           "textFormat": {"foregroundColor": fg, "bold": bold,
                                          "italic": italic}}}}}})

    red = {"red": 0.99, "green": 0.91, "blue": 0.91}
    amber = {"red": 1.0, "green": 0.95, "blue": 0.88}
    green = {"red": 0.92, "green": 0.96, "blue": 0.93}
    grey_bg = {"red": 0.97, "green": 0.97, "blue": 0.97}
    dark_red = {"red": 0.61, "green": 0.11, "blue": 0.11}
    dark_amber = {"red": 0.54, "green": 0.29, "blue": 0.0}
    dark_green = {"red": 0.11, "green": 0.37, "blue": 0.13}
    grey_fg = {"red": 0.6, "green": 0.65, "blue": 0.7}

    # What to do, loudest first.
    rule("Do next", "TEXT_EQ", ["Contact now"], red, dark_red, bold=True)
    rule("Do next", "TEXT_EQ", ["Follow up due"], amber, dark_amber, bold=True)
    rule("Do next", "TEXT_STARTS_WITH", ["Chase"], amber, dark_amber)
    rule("Do next", "TEXT_EQ", ["Closed"], grey_bg, grey_fg, italic=True)
    # Who is worth the most.
    rule("Priority", "TEXT_EQ", ["1 - Hot"], red, dark_red, bold=True)
    rule("Priority", "TEXT_EQ", ["2 - High"], amber, dark_amber, bold=True)
    rule("Status", "TEXT_EQ", ["Won"], green, dark_green, bold=True)
    rule("Status", "TEXT_EQ", ["Lost"], grey_bg, grey_fg, italic=True)
    # A lead nobody has touched in a week is what actually leaks money.
    rule("Days open", "NUMBER_GREATER_THAN_EQ", ["7"], red, dark_red, bold=True)

    google_json(token,
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}:batchUpdate",
                method="POST", payload={"requests": requests})
    print(f"leads tab: status + owner dropdowns, 9 colour rules "
          f"(replaced {existing})")


def freeze_and_widths(token, sheet_id):
    """Freeze the identity columns and re-apply every width from LEAD_COLS."""
    leads_id = tab_id(token, sheet_id, "Leads")
    frozen = ALL_LEAD_HEADERS.index("Company")
    # The title and subtitle were merged across the old width, and Sheets
    # refuses to freeze a column boundary that cuts a merged cell in half.
    # Unmerged text still spills across empty neighbours, so nothing is lost.
    requests = [{"unmergeCells": {"range": {
        "sheetId": leads_id, "startRowIndex": 0, "endRowIndex": HEADER_ROW - 1,
        "startColumnIndex": 0, "endColumnIndex": LEAD_ID_COL}}}]
    requests.append({"updateSheetProperties": {
        "properties": {"sheetId": leads_id,
                       "gridProperties": {"frozenRowCount": HEADER_ROW,
                                          "frozenColumnCount": frozen}},
        "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"}})
    for i, (label, width) in enumerate(LEAD_COLS + LEAD_EDIT_COLS):
        requests.append({"updateDimensionProperties": {
            "range": {"sheetId": leads_id, "dimension": "COLUMNS",
                      "startIndex": i, "endIndex": i + 1},
            "properties": {"pixelSize": width * 7 + 5}, "fields": "pixelSize"}})
    # The id column stays hidden; it is plumbing, not information.
    requests.append({"updateDimensionProperties": {
        "range": {"sheetId": leads_id, "dimension": "COLUMNS",
                  "startIndex": LEAD_ID_COL - 1, "endIndex": LEAD_ID_COL},
        "properties": {"hiddenByUser": True}, "fields": "hiddenByUser"}})
    google_json(token,
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}:batchUpdate",
                method="POST", payload={"requests": requests})

    google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/"
        "Leads!A1:A2?valueInputOption=RAW", method="PUT",
        payload={"range": "Leads!A1:A2", "majorDimension": "ROWS", "values": [
            [LEADS_TITLE], [LEADS_SUBTITLE]]})
    print(f"leads tab: froze {frozen} columns, re-applied widths")


def dashboard_formulas():
    """Every Dashboard label mapped to the formula it should hold.

    Derived from ALL_LEAD_HEADERS rather than left to Google's auto-adjust,
    because auto-adjust gets one case wrong: a reference to the exact column an
    insert lands on stays put instead of following its data. That silently
    zeroed the 'By track' counts the first time these columns moved.
    """
    lo, hi = FIRST_DATA_ROW, FIRST_DATA_ROW + LEAD_BUFFER_ROWS - 1

    def count(label, value):
        c = lead_col(label)
        return f'=COUNTIF(Leads!${c}${lo}:${c}${hi},"{value}")'

    em, fl = lead_col("Email"), lead_col("Flag")
    days, st = lead_col("Days open"), lead_col("Status")
    pr = lead_col("Priority")
    out = {
        # COUNTIFS, not SUMPRODUCT: an empty "Days open" cell holds "" and in
        # Sheets text always ranks above a number, so ""<=>7 would count every
        # blank row in the 400-row buffer as overdue.
        "Waiting for a first touch": count("Do next", "Contact now"),
        "Follow-up due today": count("Do next", "Follow up due"),
        "Untouched 7+ days":
            f'=COUNTIFS(Leads!${days}${lo}:${days}${hi},">=7",'
            f'Leads!${st}${lo}:${st}${hi},"New")',
        # Both live on the Leads tab now, so both counts come from it. Splitting
        # on Priority rather than Entry point keeps this true if the popup copy
        # or the source strings ever change.
        "Leads (full details)":
            f'=SUMPRODUCT((Leads!${em}${lo}:${em}${hi}<>"")'
            f'*(Leads!${pr}${lo}:${pr}${hi}<>"{EMAIL_ONLY_PRIORITY}"))',
        "Email-only signups": f'=COUNTIF(Leads!${pr}${lo}:${pr}${hi},'
                              f'"{EMAIL_ONLY_PRIORITY}")',
        # The KPI for the popup's optional second step: an email-only signup
        # that left a number stopped being email-only and became callable.
        "…of those, callable":
            f'=COUNTIFS(Leads!${pr}${lo}:${pr}${hi},"{EMAIL_ONLY_PRIORITY}",'
            f'Leads!${lead_col("Phone")}${lo}:${lead_col("Phone")}${hi},"<>")',
        "Unique lead emails":
            f'=SUMPRODUCT((Leads!${fl}${lo}:${fl}${hi}<>"earlier dup")'
            f'*(Leads!${em}${lo}:${em}${hi}<>""))',
        "Newest lead": f'=IFERROR(TEXT(MAX(Leads!$A${lo}:$A${hi}),"yyyy-mm-dd"),"-")',
        "Added optional details": extra_details_formula(lo, hi),
    }
    for p in PRIORITY_ORDER:
        out[p] = count("Priority", p)
    for s in STATUS_OPTIONS:
        out[s] = count("Status", s)
    for s in TRAFFIC_BUCKETS:
        out[s] = count("Came from", s)
    for k in ("A", "B", "C"):
        out[TRACK_LABELS[k]] = count("Track", TRACK_LABELS[k])
    for v in VOLUME_TIER_LABELS.values():
        out[v] = count("Monthly volume", v)
    return out


def dashboard_layout():
    """Blocks as (label_col, value_col, header_row, title, labels).

    Order is deliberate: the three numbers that decide what Linoy does in the
    next hour sit top-left, where the eye lands. Everything else is reporting.
    """
    return [
        ("A", "B", 4, "Work queue", ["Waiting for a first touch",
                                     "Follow-up due today", "Untouched 7+ days"]),
        ("A", "B", 9, "Totals", ["Leads (full details)", "Email-only signups",
                                 "…of those, callable", "Unique lead emails",
                                 "Added optional details", "Newest lead"]),
        ("A", "B", 16, "By status", list(STATUS_OPTIONS)),
        ("D", "E", 4, "By priority", list(PRIORITY_ORDER)),
        ("D", "E", 11, "Where they came from", list(TRAFFIC_BUCKETS)),
        ("D", "E", 19, "By track", [TRACK_LABELS[k] for k in ("A", "B", "C")]),
        ("D", "E", 24, "By monthly volume", list(VOLUME_TIER_LABELS.values())),
    ]


def rebuild_dashboard(token, sheet_id):
    """Rewrite the whole Dashboard from dashboard_layout().

    Safe to do wholesale because this tab holds nothing a human typed: it is
    labels and formulas, every one of them derived from the Leads tab.
    """
    formula = dashboard_formulas()
    grid = [["" for _ in range(5)] for _ in range(40)]
    grid[0][0] = "Overview"
    grid[1][0] = ("Live formulas, nothing to fill in. Work queue is what needs "
                  "doing today.")

    header_cells = []
    for lc, vc, start, title, labels in dashboard_layout():
        col = 0 if lc == "A" else 3
        grid[start - 1][col] = title
        header_cells.append((start, lc, vc))
        for i, label in enumerate(labels, start=start):
            grid[i][col] = label
            grid[i][col + 1] = formula[label]

    google_json(
        token,
        f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/"
        "Dashboard!A1:E40?valueInputOption=USER_ENTERED", method="PUT",
        payload={"range": "Dashboard!A1:E40", "majorDimension": "ROWS",
                 "values": grid})

    dash_id = tab_id(token, sheet_id, "Dashboard")
    # Wipe formatting first. Blocks have moved rows since this tab was built,
    # so old header styling would otherwise stay bolded onto whatever label
    # now sits at that row.
    requests = [{"repeatCell": {
        "range": {"sheetId": dash_id, "startRowIndex": 2, "endRowIndex": 40,
                  "startColumnIndex": 0, "endColumnIndex": 5},
        "cell": {"userEnteredFormat": {
            "backgroundColor": {"red": 1, "green": 1, "blue": 1},
            "horizontalAlignment": "LEFT",
            "textFormat": {"bold": False, "foregroundColor": {
                "red": 0.11, "green": 0.15, "blue": 0.20}}}},
        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,textFormat)"}}]

    # Values bold and right-aligned so the number, not the label, is what reads.
    for lc, vc, start, title, labels in dashboard_layout():
        col = 1 if lc == "A" else 4
        requests.append({"repeatCell": {
            "range": {"sheetId": dash_id, "startRowIndex": start,
                      "endRowIndex": start + len(labels),
                      "startColumnIndex": col, "endColumnIndex": col + 1},
            "cell": {"userEnteredFormat": {
                "horizontalAlignment": "RIGHT",
                "textFormat": {"bold": True, "foregroundColor": {
                    "red": 0.06, "green": 0.15, "blue": 0.25}}}},
            "fields": "userEnteredFormat(horizontalAlignment,textFormat)"}})

    for start, lc, vc in header_cells:
        col = 0 if lc == "A" else 3
        requests.append({"repeatCell": {
            "range": {"sheetId": dash_id, "startRowIndex": start - 1,
                      "endRowIndex": start, "startColumnIndex": col,
                      "endColumnIndex": col + 2},
            "cell": {"userEnteredFormat": {
                "backgroundColor": {"red": 0.96, "green": 0.97, "blue": 0.98},
                "textFormat": {"bold": True, "foregroundColor": {
                    "red": 0.06, "green": 0.15, "blue": 0.25}}}},
            "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
    google_json(token,
                f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}:batchUpdate",
                method="POST", payload={"requests": requests})
    print(f"dashboard: rebuilt {len(header_cells)} blocks, work queue on top")


def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "sync"
    env = read_env(ENV_FILE) if os.path.exists(ENV_FILE) else {}
    url = env.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get(
        "SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("missing Supabase config: need .env.local or SUPABASE_URL + "
                 "SUPABASE_SERVICE_ROLE_KEY in the environment")
    if mode == "create" and not HAVE_OPENPYXL:
        sys.exit("create mode needs openpyxl (pip install openpyxl)")

    quiz_rows = supabase_get(url, key, "quiz_leads",
                             "select=*&order=created_at.desc")
    sub_rows = supabase_get(url, key, "newsletter_subscribers",
                            "select=*&order=subscribed_at.desc")
    # One call list on the Leads tab. The Newsletter tab stays as the mailing
    # list of record (it keeps unsubscribes, which the call list drops).
    leads = build_lead_stream(quiz_rows, sub_rows)
    subs = build_newsletter(sub_rows)
    email_only = sum(1 for x in leads if x["priority"] == EMAIL_ONLY_PRIORITY)
    print(f"pulled {len(leads)} rows for Leads "
          f"({len(leads) - email_only} full, {email_only} email-only), "
          f"{len(subs)} on the newsletter list")

    if len(leads) > LEAD_BUFFER_ROWS or len(subs) > NEWS_BUFFER_ROWS:
        sys.exit("buffer exceeded — raise LEAD_BUFFER_ROWS / NEWS_BUFFER_ROWS "
                 "and re-run 'create'")

    token = google_access_token()
    state = {}
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as fh:
            state = json.load(fh)

    if mode == "migrate-cols":
        sid = state.get("spreadsheet_id")
        if not sid:
            sys.exit("no spreadsheet to migrate: missing " + STATE_FILE)
        migrate_columns(token, sid)
        freeze_and_widths(token, sid)
        refresh_leads_rules(token, sid)
        refresh_newsletter_header(token, sid)
        rebuild_dashboard(token, sid)
        rebuild_legend(token, sid)
        sync_tab(token, sid, "Leads", leads, lead_row_values,
                 N_LEAD_DATA, N_LEAD_TOTAL, LEAD_ID_COL, LEAD_BUFFER_ROWS)
        sync_tab(token, sid, "Newsletter", subs, news_row_values,
                 N_NEWS_DATA, N_NEWS_TOTAL, NEWS_ID_COL, NEWS_BUFFER_ROWS)
        print(f"migrated {state['url']}")
    elif mode == "create" or not state.get("spreadsheet_id"):
        build_workbook(leads, subs, XLSX_FILE)
        info = upload_new(token, XLSX_FILE)
        state = {"spreadsheet_id": info["id"],
                 "url": info.get("webViewLink") or
                 f"https://docs.google.com/spreadsheets/d/{info['id']}/edit"}
        with open(STATE_FILE, "w") as fh:
            json.dump(state, fh, indent=2)
        print(f"created {state['url']}")
    else:
        sid = state["spreadsheet_id"]
        sync_tab(token, sid, "Leads", leads, lead_row_values,
                 N_LEAD_DATA, N_LEAD_TOTAL, LEAD_ID_COL, LEAD_BUFFER_ROWS)
        sync_tab(token, sid, "Newsletter", subs, news_row_values,
                 N_NEWS_DATA, N_NEWS_TOTAL, NEWS_ID_COL, NEWS_BUFFER_ROWS)
        print(f"synced {state['url']}")

    print(json.dumps({"leads": len(leads), "email_only": email_only,
                      "newsletter": len(subs), "url": state.get("url")},
                     indent=2))


if __name__ == "__main__":
    main()
